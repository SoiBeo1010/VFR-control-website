import fastapi
from fastapi import FastAPI, HTTPException, Depends, status, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from datetime import datetime, timedelta
import json
import jwt
import bcrypt
from sqlalchemy import create_engine, Column, Integer, String, DateTime, Float, Text, ForeignKey
from sqlalchemy.orm import declarative_base, sessionmaker, Session
from typing import Optional, List, Dict, Any
from pydantic import BaseModel, ConfigDict
import os
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
import uvicorn

load_dotenv()

app = FastAPI(title="QMS System")

# CORS configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# SQLite Database configuration
DATABASE_URL = "sqlite:///./qms_system.db"
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

SECRET_KEY = os.getenv("SECRET_KEY", "your-secret-key-change-in-production")
ALGORITHM = "HS256"

# ============ DATABASE MODELS ============

class UserDB(Base):
    __tablename__ = "users"
    user_id = Column(Integer, primary_key=True, index=True)
    username = Column(String(255), unique=True, nullable=False, index=True)
    password = Column(String(255), nullable=False)
    role = Column(String(50), nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)

class InspectionDB(Base):
    __tablename__ = "inspections"
    inspection_id = Column(Integer, primary_key=True, index=True)
    product_id = Column(String(100), nullable=False)
    batch_number = Column(String(100), nullable=False)
    wo_number = Column(String(100))
    component_code = Column(String(100))
    carcass_code = Column(String(100))
    inspection_stage = Column(String(100))  # Finishing, Sanding, Final, Assembly, Item White, Machinery, Upholstery
    inspector_id = Column(Integer, ForeignKey("users.user_id"), nullable=False)
    inspection_date = Column(DateTime, default=datetime.utcnow)
    qty_inspected = Column(Integer, default=1)
    qty_passed = Column(Integer, default=0)
    qty_failed = Column(Integer, default=0)
    status = Column(String(50), default="pending")
    repair_status = Column(String(50), nullable=True)
    repair_status_updated_at = Column(DateTime, nullable=True)

class DefectListDB(Base):
    __tablename__ = "defect_list"
    defect_id = Column(Integer, primary_key=True, index=True)
    inspection_id = Column(Integer, ForeignKey("inspections.inspection_id"), nullable=False)
    defect_code = Column(String(100), nullable=False)
    defect_type = Column(String(100), nullable=False)
    defect_category = Column(String(50))  # Minor, Major
    defect_description = Column(Text)
    severity = Column(String(50))
    material_standard = Column(String(100))
    inspection_method = Column(String(100))  # Appearance, Measurement, Template/Drawing
    remark = Column(Text)
    created_date = Column(DateTime, default=datetime.utcnow)

class QuarantineReportDB(Base):
    __tablename__ = "quarantine_report"
    quarantine_id = Column(Integer, primary_key=True, index=True)
    inspection_id = Column(Integer, ForeignKey("inspections.inspection_id"), nullable=False)
    defect_list_id = Column(Integer, ForeignKey("defect_list.defect_id"), nullable=False)
    created_date = Column(DateTime, default=datetime.utcnow)

class OwnerAssignmentDB(Base):
    __tablename__ = "owner_assignment"
    assignment_id = Column(Integer, primary_key=True, index=True)
    quarantine_id = Column(Integer, ForeignKey("quarantine_report.quarantine_id"), nullable=False)
    problem_owner_id = Column(Integer, ForeignKey("users.user_id"), nullable=False)
    repair_department_id = Column(Integer)
    priority = Column(String(50))
    assigned_date = Column(DateTime, default=datetime.utcnow)

class RepairPlanDB(Base):
    __tablename__ = "repair_plan"
    plan_id = Column(Integer, primary_key=True, index=True)
    assignment_id = Column(Integer, ForeignKey("owner_assignment.assignment_id"), nullable=False)
    plan_description = Column(Text)
    estimated_days = Column(Integer)
    assigned_worker_id = Column(Integer, ForeignKey("users.user_id"))
    created_date = Column(DateTime, default=datetime.utcnow)
    status = Column(String(50), default="pending")

class ApprovalDB(Base):
    __tablename__ = "approval"
    approval_id = Column(Integer, primary_key=True, index=True)
    repair_plan_id = Column(Integer, ForeignKey("repair_plan.plan_id"), nullable=False)
    status = Column(String(50))
    approver_id = Column(Integer, ForeignKey("users.user_id"))
    comments = Column(Text)
    approval_date = Column(DateTime, default=datetime.utcnow)

class RepairCompletionDB(Base):
    __tablename__ = "repair_completion"
    completion_id = Column(Integer, primary_key=True, index=True)
    repair_plan_id = Column(Integer, ForeignKey("repair_plan.plan_id"), nullable=False)
    labor_cost = Column(Float, default=0)
    spare_cost = Column(Float, default=0)
    actual_cost = Column(Float)
    completion_date = Column(DateTime)
    root_cause = Column(Text)
    preventive_action = Column(Text)

class CostManagementDB(Base):
    __tablename__ = "cost_management"
    cost_id = Column(Integer, primary_key=True, index=True)
    repair_completion_id = Column(Integer, ForeignKey("repair_completion.completion_id"), nullable=False)
    cost_type = Column(String(100))
    amount = Column(Float)
    recorded_date = Column(DateTime, default=datetime.utcnow)

class RepairCardSupplementLinkDB(Base):
    __tablename__ = "repair_card_supplement_link"
    link_id = Column(Integer, primary_key=True, index=True)
    repair_card_id = Column(String(100), index=True, nullable=False)
    supplement_request_code = Column(String(100), index=True, nullable=False)
    wo_no = Column(String(100), nullable=False)
    item_code = Column(String(100), nullable=False)
    carcass_code = Column(String(100), nullable=True)
    linked_at = Column(DateTime, default=datetime.utcnow)

# Create all tables
Base.metadata.create_all(bind=engine)

# Auto-migration script for SQLite to add missing columns in inspections table
try:
    with engine.connect() as conn:
        from sqlalchemy import text
        cursor = conn.execute(text("PRAGMA table_info(inspections)"))
        columns = [row[1] for row in cursor.fetchall()]
        if "repair_status" not in columns:
            conn.execute(text("ALTER TABLE inspections ADD COLUMN repair_status VARCHAR(50)"))
            print("Migration: Added column 'repair_status' to inspections table.")
        if "repair_status_updated_at" not in columns:
            conn.execute(text("ALTER TABLE inspections ADD COLUMN repair_status_updated_at DATETIME"))
            print("Migration: Added column 'repair_status_updated_at' to inspections table.")
        conn.commit()
except Exception as migration_err:
    print(f"Auto-migration warning: {migration_err}")

# Dependency
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ============ PYDANTIC MODELS ============

class User(BaseModel):
    username: str
    password: str
    role: str

class LoginRequest(BaseModel):
    username: str
    password: str

class LoginResponse(BaseModel):
    token: str
    user_id: int
    username: str
    role: str

class Inspection(BaseModel):
    inspection_id: Optional[int] = None
    product_id: str
    batch_number: str
    inspector_id: int
    inspection_date: Optional[str] = None
    status: str = "pending"

class DefectItem(BaseModel):
    defect_id: Optional[int] = None
    inspection_id: int
    defect_type: str
    defect_description: str
    severity: str

class QuarantineReport(BaseModel):
    quarantine_id: Optional[int] = None
    inspection_id: int
    defect_list_id: int
    created_date: Optional[str] = None

class OwnerAssignment(BaseModel):
    assignment_id: Optional[int] = None
    quarantine_id: int
    problem_owner_id: int
    repair_department_id: int
    priority: str

class RepairPlan(BaseModel):
    plan_id: Optional[int] = None
    assignment_id: int
    plan_description: str
    estimated_days: int
    assigned_worker_id: int

class Approval(BaseModel):
    approval_id: Optional[int] = None
    repair_plan_id: int
    status: str
    approver_id: int
    comments: Optional[str] = None

class RepairCompletion(BaseModel):
    completion_id: Optional[int] = None
    repair_plan_id: int
    actual_cost: float
    completion_date: Optional[str] = None
    root_cause: str
    preventive_action: str


INSPECTION_DATA_FILE = Path(__file__).resolve().parent / "data" / "inspection-control.json"
INSPECTION_FIELDS = [
    "branch_itw",
    "branch_fgw",
    "wo_no",
    "item_code",
    "partcode",
    "wo_quantity",
    "inspector",
    "inspector_name",
    "inspector_dept",
    "inspection_stage",
    "inspected_date",
    "inspected_time",
    "inspected_qty",
    "pass_qty",
    "fail_qty",
    "quarantine",
    "defect_code",
    "defect_category",
    "defect_owner",
    "inspector_recommend",
    "repair_department",
    "remark",
    "check_pass_date",
    "check_pass_minutes",
    "check_pass_hours",
]

# ============ UTILITY FUNCTIONS ============

def hash_password(password: str) -> str:
    """Hash password using bcrypt"""
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(plain_password: str, hashed_password: str) -> bool:
    """Verify password against hash"""
    return bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    """Create JWT token"""
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(hours=24)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def verify_token(token: str):
    """Verify JWT token"""
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        user_id: int = payload.get("user_id")
        role: str = payload.get("role")
        if username is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        return {"username": username, "user_id": user_id, "role": role}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


def _coerce_int(value: Any, default: int = 0) -> int:
    try:
        if value in (None, "", "None"):
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _coerce_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value in (None, "", 0, "0", "false", "False", "no", "No"):
        return False
    return True


def _serialize_datetime(value: Any) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    try:
        parsed = datetime.fromisoformat(str(value))
        return parsed.isoformat()
    except ValueError:
        return str(value)


def _parse_datetime(value: Any) -> Optional[datetime]:
    if value in (None, "", "None"):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    try:
        dt = datetime.fromisoformat(str(value))
        return dt.replace(tzinfo=None)
    except ValueError:
        return None


def _load_inspection_records() -> List[Dict[str, Any]]:
    if not INSPECTION_DATA_FILE.exists():
        return []
    try:
        return json.loads(INSPECTION_DATA_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []


def _save_inspection_records(records: List[Dict[str, Any]]) -> None:
    INSPECTION_DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
    INSPECTION_DATA_FILE.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")


def _resolve_qc_dashboard_source_file(source_path: Optional[str] = None) -> Path:
    candidate = source_path or os.getenv("QC_DASHBOARD_SOURCE_PATH")
    if not candidate:
        raise FileNotFoundError("QC_DASHBOARD_SOURCE_PATH environment variable is not defined.")
    candidate_path = Path(candidate).expanduser()
    if not candidate_path.is_absolute():
        candidate_path = (Path(__file__).resolve().parent / candidate_path).resolve()
    return candidate_path


def _excel_cell_to_iso_datetime(value: Any) -> Optional[str]:
    if value in (None, "", "None"):
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).isoformat()
        except Exception:
            return None
    try:
        parsed = datetime.fromisoformat(str(value))
        return parsed.isoformat()
    except ValueError:
        try:
            parsed = datetime.strptime(str(value), "%m/%d/%Y")
            return parsed.isoformat()
        except ValueError:
            return str(value)


def _excel_cell_to_int(value: Any, default: int = 0) -> int:
    try:
        if value in (None, "", "None"):
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _excel_cell_to_optional_int(value: Any) -> Optional[int]:
    if value in (None, "", "None"):
        return None
    return 1 if _coerce_bool(value) else 0


def _excel_cell_to_optional_text(value: Any) -> Optional[str]:
    if value in (None, "", "None"):
        return None
    text = str(value).strip()
    return text or None


# Cache for QC Dashboard rows
_qc_dashboard_cache = {
    "source_path": None,
    "mtime": 0.0,
    "rows": []
}

def _load_qc_dashboard_source_rows(source_path: Optional[str] = None) -> List[Dict[str, Any]]:
    global _qc_dashboard_cache
    workbook_path = _resolve_qc_dashboard_source_file(source_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Dashboard source file not found: {workbook_path}")

    try:
        current_mtime = os.path.getmtime(workbook_path)
    except Exception:
        current_mtime = 0.0

    # Return cached data if path and modification time match
    if (
        _qc_dashboard_cache["source_path"] == str(workbook_path)
        and _qc_dashboard_cache["mtime"] == current_mtime
        and _qc_dashboard_cache["rows"]
    ):
        return _qc_dashboard_cache["rows"]

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows_iter = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []

    headers = [str(header).strip() if header is not None else "" for header in header_row]
    header_indexes = {header: index for index, header in enumerate(headers) if header}

    def get_value(row_values: tuple[Any, ...], column_name: str) -> Any:
        index = header_indexes.get(column_name)
        if index is None or index >= len(row_values):
            return None
        return row_values[index]

    normalized_rows: List[Dict[str, Any]] = []
    for row_values in rows_iter:
        normalized_rows.append({
            "inspectorName": _excel_cell_to_optional_text(get_value(row_values, "Inspector Name")) or "",
            "inspectionStage": _excel_cell_to_optional_text(get_value(row_values, "Inspection Stage")) or "",
            "inspectedDate": _excel_cell_to_iso_datetime(get_value(row_values, "Inspected Date"))
            or _excel_cell_to_iso_datetime(get_value(row_values, "Inspected Time"))
            or "",
            "inspectedQty": _excel_cell_to_int(get_value(row_values, "Inspected Qty"), 0),
            "passQty": _excel_cell_to_int(get_value(row_values, "Pass Qty"), 0),
            "failQty": _excel_cell_to_int(get_value(row_values, "Fail Qty"), 0),
            "quarantine": _excel_cell_to_optional_int(get_value(row_values, "Quarantine")),
            "defectCode": _excel_cell_to_optional_text(get_value(row_values, "Defect Code")),
            "defectOwner": _excel_cell_to_optional_text(get_value(row_values, "Defect Owner")),
            "branchITW": _excel_cell_to_optional_text(get_value(row_values, "Branch ITW")) or "",
            "inspectorDept": _excel_cell_to_optional_text(get_value(row_values, "Inspector Dept.")) or "",
        })

    _qc_dashboard_cache = {
        "source_path": str(workbook_path),
        "mtime": current_mtime,
        "rows": normalized_rows
    }
    return normalized_rows



def _record_status(record: Dict[str, Any]) -> str:
    status = str(record.get("status") or "").strip()
    if status:
        normalized = status.lower()
        if normalized == "pending":
            return "Pending"
        if normalized in {"pass", "fail"}:
            return normalized.capitalize()
        return status
    if _coerce_bool(record.get("quarantine")) or _coerce_int(record.get("fail_qty")) > 0:
        return "Fail"
    return "Pass"


def _normalize_inspection_record(record: Dict[str, Any], record_id: Optional[int] = None, source_file: str = "QC Worker", source_row: Optional[int] = None) -> Dict[str, Any]:
    normalized = {field: record.get(field) for field in INSPECTION_FIELDS}
    normalized["branch_itw"] = normalized["branch_itw"] or ""
    normalized["branch_fgw"] = normalized["branch_fgw"] or ""
    normalized["wo_no"] = normalized["wo_no"] or record.get("wo_number") or record.get("wo") or record.get("product_id") or ""
    normalized["item_code"] = normalized["item_code"] or record.get("itemCode") or record.get("component_code") or record.get("batch_number") or ""
    normalized["partcode"] = normalized["partcode"] or record.get("partCode") or record.get("carcass_code") or ""
    normalized["wo_quantity"] = _coerce_int(normalized["wo_quantity"], 0)
    normalized["inspector"] = normalized["inspector"] or record.get("inspector_id") or ""
    normalized["inspector_name"] = normalized["inspector_name"] or record.get("inspectorName") or record.get("inspector_name") or ""
    normalized["inspector_dept"] = normalized["inspector_dept"] or record.get("inspectorDept") or record.get("inspector_dept") or ""
    normalized["inspection_stage"] = normalized["inspection_stage"] or record.get("inspectionStage") or record.get("inspection_stage") or ""
    normalized["inspected_date"] = _serialize_datetime(normalized["inspected_date"] or record.get("inspection_date"))
    normalized["inspected_time"] = _serialize_datetime(normalized["inspected_time"] or record.get("inspection_date"))
    normalized["inspected_qty"] = _coerce_int(normalized["inspected_qty"], 0)
    normalized["pass_qty"] = _coerce_int(normalized["pass_qty"], 0)
    normalized["fail_qty"] = _coerce_int(normalized["fail_qty"], 0)
    normalized["quarantine"] = 1 if _coerce_bool(normalized["quarantine"] or record.get("quarantine")) else 0
    normalized["defect_code"] = normalized["defect_code"] or record.get("defectCode") or ""
    normalized["defect_category"] = normalized["defect_category"] or record.get("defectCategory") or ""
    normalized["defect_owner"] = normalized["defect_owner"] or record.get("defectOwner") or ""
    normalized["inspector_recommend"] = normalized["inspector_recommend"] or record.get("inspectorRecommend") or ""
    normalized["repair_department"] = normalized["repair_department"] or record.get("repairDepartment") or ""
    normalized["remark"] = normalized["remark"] or record.get("remark") or ""
    normalized["check_pass_date"] = _serialize_datetime(normalized["check_pass_date"])
    normalized["check_pass_minutes"] = normalized["check_pass_minutes"] if normalized["check_pass_minutes"] not in (None, "") else ""
    normalized["check_pass_hours"] = normalized["check_pass_hours"] if normalized["check_pass_hours"] not in (None, "") else ""
    normalized["id"] = _coerce_int(record_id if record_id is not None else record.get("id"), 0)
    normalized["source_file"] = source_file
    normalized["source_row"] = source_row
    normalized["status"] = _record_status({**record, **normalized})
    return normalized


def _load_store() -> List[Dict[str, Any]]:
    records = [
        _normalize_inspection_record(record, record.get("id"), record.get("source_file") or "Imported", record.get("source_row"))
        for record in _load_inspection_records()
    ]
    records.sort(key=lambda item: (_parse_datetime(item.get("inspected_date")) or datetime.min), reverse=True)
    return records


def _filter_records(
    records: List[Dict[str, Any]],
    status: Optional[str] = None,
    stage: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    department: Optional[str] = None,
    defect_code: Optional[str] = None,
) -> List[Dict[str, Any]]:
    filtered = records
    if status:
        normalized_status = status.lower()
        if normalized_status == "pass":
            filtered = [record for record in filtered if record.get("status") == "Pass" and _coerce_int(record.get("fail_qty")) == 0]
        elif normalized_status == "fail":
            filtered = [record for record in filtered if record.get("status") == "Fail" or _coerce_bool(record.get("quarantine")) or _coerce_int(record.get("fail_qty")) > 0]
        elif normalized_status == "pending":
            filtered = [record for record in filtered if str(record.get("status")).lower() == "pending"]
    if stage:
        filtered = [record for record in filtered if record.get("inspection_stage") == stage]
    if department:
        filtered = [record for record in filtered if record.get("inspector_dept") == department or record.get("repair_department") == department]
    if defect_code:
        filtered = [record for record in filtered if record.get("defect_code") == defect_code]
    if date_from:
        from_date = _parse_datetime(date_from)
        if from_date:
            filtered = [record for record in filtered if (_parse_datetime(record.get("inspected_date")) or datetime.min) >= from_date]
    if date_to:
        to_date = _parse_datetime(date_to)
        if to_date:
            filtered = [record for record in filtered if (_parse_datetime(record.get("inspected_date")) or datetime.min) <= to_date]
    return filtered


def _to_inspection_control_row(record: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": record.get("id"),
        "branch_itw": record.get("branch_itw") or "N/A",
        "branch_fgw": record.get("branch_fgw") or "N/A",
        "wo_no": record.get("wo_no") or "N/A",
        "item_code": record.get("item_code") or "N/A",
        "partcode": record.get("partcode") or "N/A",
        "wo_quantity": record.get("wo_quantity") or 0,
        "inspector": record.get("inspector") or "N/A",
        "inspector_name": record.get("inspector_name") or "N/A",
        "inspector_dept": record.get("inspector_dept") or "N/A",
        "inspection_stage": record.get("inspection_stage") or "N/A",
        "inspected_date": record.get("inspected_date") or "N/A",
        "inspected_time": record.get("inspected_time") or "N/A",
        "inspected_qty": record.get("inspected_qty") or 0,
        "pass_qty": record.get("pass_qty") or 0,
        "fail_qty": record.get("fail_qty") or 0,
        "quarantine": record.get("quarantine") or 0,
        "defect_code": record.get("defect_code") or "N/A",
        "defect_category": record.get("defect_category") or "N/A",
        "defect_owner": record.get("defect_owner") or "N/A",
        "inspector_recommend": record.get("inspector_recommend") or "N/A",
        "repair_department": record.get("repair_department") or "N/A",
        "remark": record.get("remark") or "",
        "check_pass_date": record.get("check_pass_date") or "",
        "check_pass_minutes": record.get("check_pass_minutes") or "",
        "check_pass_hours": record.get("check_pass_hours") or "",
        "status": record.get("status") or "Pass",
    }


def _match_query_value(record_value: Any, query_value: str) -> bool:
    if query_value == "":
        return True
    value_text = "" if record_value is None else str(record_value)
    return query_value.lower() in value_text.lower()


def _apply_column_filters(records: List[Dict[str, Any]], query_params: Dict[str, str], excluded_keys: set[str]) -> List[Dict[str, Any]]:
    search_query = query_params.get("search", "").strip().lower()
    filters = {key: value.strip() for key, value in query_params.items() if key not in excluded_keys and key != "search" and value is not None and str(value).strip() != ""}
    
    filtered = records
    
    # Apply global search query
    if search_query:
        searchable_fields = ["wo_no", "item_code", "partcode", "inspector_name", "defect_code", "defect_owner", "remark"]
        filtered_search = []
        for record in filtered:
            match = False
            for field in searchable_fields:
                val = record.get(field)
                if val is not None and search_query in str(val).lower():
                    match = True
                    break
            if match:
                filtered_search.append(record)
        filtered = filtered_search

    for key, value in filters.items():
        filtered = [record for record in filtered if _match_query_value(record.get(key), value)]
    return filtered


def _compute_dashboard_payload(records: List[Dict[str, Any]], include_unknown: bool = False) -> Dict[str, Any]:
    process_stages = ['Assembly', 'Final', 'Finishing', 'Sanding', 'Item White', 'Machinery', 'Upholstery']
    relevant_records = records if include_unknown else [record for record in records if record.get('inspection_stage') in process_stages]

    total_inspected = sum(_coerce_int(record.get('inspected_qty'), 1) for record in relevant_records)
    total_failed = sum(_coerce_int(record.get('fail_qty'), 0) for record in relevant_records)
    overall_fail_rate = round((total_failed / total_inspected * 100) if total_inspected > 0 else 0, 2)

    status_breakdown = {'Pass': 0, 'Fail': 0, 'Pending': 0}
    for record in relevant_records:
        if str(record.get('status') or '').lower() == 'pending':
            status_breakdown['Pending'] += 1
        elif _coerce_bool(record.get('quarantine')) or _coerce_int(record.get('fail_qty'), 0) > 0:
            status_breakdown['Fail'] += 1
        else:
            status_breakdown['Pass'] += 1

    fail_by_stage: Dict[str, Dict[str, Any]] = {}
    for stage in process_stages:
        stage_records = [record for record in relevant_records if record.get('inspection_stage') == stage]
        if not stage_records:
            continue
        stage_inspected = sum(_coerce_int(record.get('inspected_qty'), 1) for record in stage_records)
        stage_failed = sum(_coerce_int(record.get('fail_qty'), 0) for record in stage_records)
        fail_by_stage[stage] = {
            'fail_rate': round((stage_failed / stage_inspected * 100) if stage_inspected > 0 else 0, 2),
            'inspected': stage_inspected,
            'failed': stage_failed,
        }

    monthly_trends: Dict[str, Dict[str, int]] = {}
    for record in relevant_records:
        inspected_date = _parse_datetime(record.get('inspected_date')) or _parse_datetime(record.get('inspected_time'))
        if not inspected_date:
            continue
        month = inspected_date.strftime('%B')
        monthly_trends.setdefault(month, {'inspected': 0, 'failed': 0})
        monthly_trends[month]['inspected'] += _coerce_int(record.get('inspected_qty'), 1)
        monthly_trends[month]['failed'] += _coerce_int(record.get('fail_qty'), 0)

    defect_counts: Dict[str, int] = {}
    for record in relevant_records:
        if not (_coerce_bool(record.get('quarantine')) or _coerce_int(record.get('fail_qty'), 0) > 0):
            continue
        code = str(record.get('defect_code') or 'Unknown').strip() or 'Unknown'
        defect_counts[code] = defect_counts.get(code, 0) + 1

    top_defects = sorted(defect_counts.items(), key=lambda item: item[1], reverse=True)[:10]

    return {
        'overall_fail_rate': overall_fail_rate,
        'total_inspected': total_inspected,
        'total_failed': total_failed,
        'total_quarantines': sum(1 for record in relevant_records if _coerce_bool(record.get('quarantine')) or _coerce_int(record.get('fail_qty'), 0) > 0),
        'total_assignments': sum(1 for record in relevant_records if str(record.get('repair_department') or '').strip()),
        'status_breakdown': status_breakdown,
        'fail_by_stage': fail_by_stage,
        'monthly_trends': monthly_trends,
        'top_defects': [{'code': code, 'count': count} for code, count in top_defects],
    }

# Dependency for getting current user from token
def get_current_user(token: str = fastapi.Header(None)):
    if not token:
        raise HTTPException(status_code=401, detail="No token provided")
    return verify_token(token)


@app.get("/api/qc-manager/dashboard-source")
def get_qc_dashboard_source(
    current_user: dict = Depends(get_current_user),
    source_path: Optional[str] = None,
):
    """Return normalized QC dashboard rows from the configured Excel workbook."""
    if current_user['role'] != 'QC Manager':
        raise HTTPException(status_code=403, detail="Unauthorized")

    try:
        workbook_path = _resolve_qc_dashboard_source_file(source_path)
        rows = _load_qc_dashboard_source_rows(source_path)
        return {
            "source_path": str(workbook_path),
            "row_count": len(rows),
            "rows": rows,
        }
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============ INITIALIZATION ENDPOINTS ============

@app.get("/api/init-db")
def initialize_database(db: Session = Depends(get_db)):
    """Initialize database with schema and default users"""
    try:
        # Migrate existing users from 'Higher Department' to 'Warehouse' and 'higher_dept_1' to 'warehouse_1'
        from sqlalchemy import text
        try:
            db.execute(text("UPDATE users SET role = 'Warehouse' WHERE role = 'Higher Department'"))
            db.execute(text("UPDATE users SET username = 'warehouse_1' WHERE username = 'higher_dept_1'"))
            db.commit()
        except Exception as migrate_err:
            db.rollback()
            print(f"Migration error: {migrate_err}")

        # Create default users for each role
        default_users = [
            ("qc_worker_1", "password123", "QC Worker"),
            ("qc_manager_1", "password123", "QC Manager"),
            ("prod_manager_1", "password123", "Production Manager"),
            ("prod_worker_1", "password123", "Production Worker"),
            ("warehouse_1", "password123", "Warehouse")
        ]
        
        for username, password, role in default_users:
            existing_user = db.query(UserDB).filter(UserDB.username == username).first()
            if not existing_user:
                hashed_pw = hash_password(password)
                user = UserDB(username=username, password=hashed_pw, role=role)
                db.add(user)
            else:
                if existing_user.role != role:
                    existing_user.role = role
                    db.add(existing_user)
        
        db.commit()
        
        return {"message": "Database initialized successfully", "default_users": default_users}
    
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# ============ HEALTH CHECK ============

@app.get("/api/health")
def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "message": "QMS System is running"}

# ============ AUTH ENDPOINTS ============

@app.post("/api/login", response_model=LoginResponse)
def login(credentials: LoginRequest, db: Session = Depends(get_db)):
    """User login endpoint"""
    try:
        user = db.query(UserDB).filter(UserDB.username == credentials.username).first()
        
        if not user or not verify_password(credentials.password, user.password):
            raise HTTPException(status_code=401, detail="Invalid credentials")
        
        access_token = create_access_token(
            data={"sub": user.username, "user_id": user.user_id, "role": user.role}
        )
        
        return LoginResponse(
            token=access_token,
            user_id=user.user_id,
            username=user.username,
            role=user.role
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============ QC WORKER ENDPOINTS ============

@app.post("/api/qc-worker/inspection")
def create_inspection(inspection: dict, current_user: dict = Depends(get_current_user)):
    """Create a new inspection record in the shared inspection-control store."""
    if current_user['role'] != 'QC Worker':
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        records = _load_store()
        next_id = (max((record.get("id") or 0) for record in records) + 1) if records else 1
        now = datetime.utcnow().isoformat()
        payload = {
            "branch_itw": inspection.get("branch_itw") or inspection.get("branchItw") or "",
            "branch_fgw": inspection.get("branch_fgw") or inspection.get("branchFgw") or "",
            "wo_no": inspection.get("wo_no") or inspection.get("product_id") or inspection.get("wo") or "",
            "item_code": inspection.get("item_code") or inspection.get("itemCode") or inspection.get("batch_number") or "",
            "partcode": inspection.get("partcode") or inspection.get("partCode") or inspection.get("carcass_code") or "",
            "wo_quantity": _coerce_int(inspection.get("wo_quantity") or inspection.get("qty_default")),
            "inspector": current_user["user_id"],
            "inspector_name": inspection.get("inspector_name") or inspection.get("inspector") or current_user.get("username", ""),
            "inspector_dept": inspection.get("inspector_dept") or inspection.get("inspectorDept") or "QC",
            "inspection_stage": inspection.get("inspection_stage") or inspection.get("stage") or "QC Worker",
            "inspected_date": inspection.get("inspected_date") or now,
            "inspected_time": inspection.get("inspected_time") or now,
            "inspected_qty": _coerce_int(inspection.get("inspected_qty") or inspection.get("quantity_checked") or inspection.get("qty_check"), 1),
            "pass_qty": _coerce_int(inspection.get("pass_qty"), 0),
            "fail_qty": _coerce_int(inspection.get("fail_qty") or inspection.get("qty_failed"), 0),
            "quarantine": 1 if _coerce_bool(inspection.get("quarantine") or inspection.get("quarantine_reason")) else 0,
            "defect_code": inspection.get("defect_code") or "",
            "defect_category": inspection.get("defect_category") or "",
            "defect_owner": inspection.get("defect_owner") or "",
            "inspector_recommend": inspection.get("inspector_recommend") or inspection.get("final_remark") or "",
            "repair_department": inspection.get("repair_department") or "",
            "remark": inspection.get("remark") or inspection.get("final_remark") or inspection.get("quarantine_reason") or "",
            "check_pass_date": inspection.get("check_pass_date") or "",
            "check_pass_minutes": inspection.get("check_pass_minutes") or "",
            "check_pass_hours": inspection.get("check_pass_hours") or "",
            "status": "Fail" if _coerce_bool(inspection.get("quarantine")) or _coerce_int(inspection.get("fail_qty")) > 0 else "Pass",
        }
        record = _normalize_inspection_record(payload, next_id, "QC Worker", None)
        records.append(record)
        _save_inspection_records(records)

        return {"inspection_id": next_id, "status": "created", "record": record}
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/qc-worker/defect-list")
def create_defect_list(defect: DefectItem, current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Create defect list item (QC Worker)"""
    if current_user['role'] != 'QC Worker':
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        new_defect = DefectListDB(
            inspection_id=defect.inspection_id,
            defect_type=defect.defect_type,
            defect_description=defect.defect_description,
            severity=defect.severity
        )
        db.add(new_defect)
        db.commit()
        db.refresh(new_defect)
        
        return {"defect_id": new_defect.defect_id, "status": "created"}
    
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/qc-worker/quarantine")
def create_quarantine_report(quarantine: QuarantineReport, current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Create quarantine report (QC Worker)"""
    if current_user['role'] != 'QC Worker':
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        new_quarantine = QuarantineReportDB(
            inspection_id=quarantine.inspection_id,
            defect_list_id=quarantine.defect_list_id
        )
        db.add(new_quarantine)
        db.commit()
        
        # Update inspection status
        inspection = db.query(InspectionDB).filter(InspectionDB.inspection_id == quarantine.inspection_id).first()
        if inspection:
            inspection.status = "quarantined"
            db.commit()
        
        db.refresh(new_quarantine)
        
        return {"quarantine_id": new_quarantine.quarantine_id, "status": "created"}
    
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/qc-worker/inspections")
def get_inspections(current_user: dict = Depends(get_current_user)):
    """Get inspections from the shared inspection-control store."""
    if current_user['role'] != 'QC Worker':
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        records = [record for record in _load_store() if str(record.get("inspector") or "") == str(current_user["user_id"]) or record.get("inspector_name") == current_user.get("username")]
        return {"inspections": [
            {
                "inspection_id": record.get("id"),
                "product_id": record.get("wo_no"),
                "batch_number": record.get("item_code"),
                "status": str(record.get("status") or "Pass").lower(),
                "inspection_date": record.get("inspected_date") or record.get("inspected_time"),
            } for record in records
        ]}
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============ QC MANAGER ENDPOINTS ============

@app.get("/api/qc-manager/quarantine-reports")
def get_quarantine_reports(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get all quarantine reports (QC Manager)"""
    if current_user['role'] != 'QC Manager':
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        reports = db.query(QuarantineReportDB).all()
        result = []
        
        for qr in reports:
            inspection = db.query(InspectionDB).filter(InspectionDB.inspection_id == qr.inspection_id).first()
            defect = db.query(DefectListDB).filter(DefectListDB.defect_id == qr.defect_list_id).first()
            assignment = db.query(OwnerAssignmentDB).filter(OwnerAssignmentDB.quarantine_id == qr.quarantine_id).first()
            inspector = db.query(UserDB).filter(UserDB.user_id == inspection.inspector_id).first() if inspection else None
            
            result.append({
                "quarantine_id": qr.quarantine_id,
                "product_id": inspection.product_id if inspection else "N/A",
                "batch_number": inspection.batch_number if inspection else "N/A",
                "wo_number": inspection.wo_number if inspection else None,
                "component_code": inspection.component_code if inspection else None,
                "carcass_code": inspection.carcass_code if inspection else None,
                "inspection_stage": inspection.inspection_stage if inspection else None,
                "inspector_name": inspector.username if inspector else "N/A",
                "inspection_date": inspection.inspection_date.isoformat() if inspection and inspection.inspection_date else None,
                "defect_type": defect.defect_type if defect else "N/A",
                "defect_code": defect.defect_code if defect else None,
                "severity": defect.severity if defect else "N/A",
                "defect_description": defect.defect_description if defect else None,
                "material_standard": defect.material_standard if defect else None,
                "inspection_method": defect.inspection_method if defect else None,
                "remark": defect.remark if defect else None,
                "assignment_status": "assigned" if assignment else "unassigned",
                "problem_owner_id": assignment.problem_owner_id if assignment else None,
                "repair_department_id": assignment.repair_department_id if assignment else None,
                "priority": assignment.priority if assignment else None,
                "created_date": qr.created_date.isoformat() if qr.created_date else None
            })
        
        return {"reports": result}
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/qc-manager/owner-assignments")
def get_owner_assignments(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get all owner assignments for QC Manager"""
    if current_user['role'] != 'QC Manager':
        raise HTTPException(status_code=403, detail="Unauthorized")

    try:
        assignments = db.query(OwnerAssignmentDB).all()
        result = []

        for assignment in assignments:
            quarantine = db.query(QuarantineReportDB).filter(QuarantineReportDB.quarantine_id == assignment.quarantine_id).first()
            inspection = db.query(InspectionDB).filter(InspectionDB.inspection_id == quarantine.inspection_id).first() if quarantine else None
            owner = db.query(UserDB).filter(UserDB.user_id == assignment.problem_owner_id).first()

            result.append({
                "assignment_id": assignment.assignment_id,
                "quarantine_id": assignment.quarantine_id,
                "product_id": inspection.product_id if inspection else "N/A",
                "batch_number": inspection.batch_number if inspection else "N/A",
                "inspection_stage": inspection.inspection_stage if inspection else None,
                "problem_owner_id": assignment.problem_owner_id,
                "problem_owner_name": owner.username if owner else "N/A",
                "repair_department_id": assignment.repair_department_id,
                "priority": assignment.priority,
                "assigned_date": assignment.assigned_date.isoformat() if assignment.assigned_date else None
            })

        return {"assignments": result}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/qc-manager/owner-assignment")
def assign_owner(assignment: OwnerAssignment, current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Assign problem owner and repair department (QC Manager)"""
    if current_user['role'] != 'QC Manager':
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        new_assignment = OwnerAssignmentDB(
            quarantine_id=assignment.quarantine_id,
            problem_owner_id=assignment.problem_owner_id,
            repair_department_id=assignment.repair_department_id,
            priority=assignment.priority
        )
        db.add(new_assignment)
        db.commit()
        db.refresh(new_assignment)
        
        return {"assignment_id": new_assignment.assignment_id, "status": "assigned"}
    
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/qc-manager/notifications")
def get_notifications(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get notifications for QC Manager"""
    if current_user['role'] != 'QC Manager':
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        inspections = db.query(InspectionDB).all()
        notifications = []
        
        for inspection in inspections:
            defects = db.query(DefectListDB).filter(DefectListDB.inspection_id == inspection.inspection_id).all()
            if defects:
                notifications.append({
                    "notification_id": inspection.inspection_id,
                    "product_id": inspection.product_id,
                    "batch_number": inspection.batch_number,
                    "inspection_stage": inspection.inspection_stage,
                    "status": "Pass" if inspection.qty_failed == 0 else "Fail",
                    "defect_count": len(defects),
                    "severity_count": len([d for d in defects if d.severity]),
                    "message": f"New inspection result for {inspection.product_id}",
                    "created_date": inspection.inspection_date.isoformat() if inspection.inspection_date else None
                })
        
        return {"notifications": notifications}
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/qc-manager/send-notification")
def send_notification(data: dict, current_user: dict = Depends(get_current_user)):
    """Send notification to Production Manager (QC Manager)"""
    if current_user['role'] != 'QC Manager':
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        return {"status": "notification_sent", "message": f"Notification sent to Production Manager"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/qc-manager/fail-file-report")
def get_fail_file_report(current_user: dict = Depends(get_current_user)):
    """Get the imported fail-file report for QC Manager"""
    if current_user['role'] != 'QC Manager':
        raise HTTPException(status_code=403, detail="Unauthorized")

    data_file = Path(__file__).resolve().parent / "public" / "inspection-data.json"
    if not data_file.exists():
        raise HTTPException(status_code=404, detail="inspection-data.json not found")

    try:
        tickets = json.loads(data_file.read_text(encoding="utf-8"))
        rows = []
        pass_count = 0
        fail_count = 0
        high_priority_count = 0

        for ticket in tickets:
            source = ticket.get("source") or {}
            fail_qty = int(source.get("failQty") or 0)
            pass_qty = int(source.get("passQty") or 0)
            priority = ticket.get("priority") or "N/A"

            if fail_qty > 0:
                fail_count += 1
            else:
                pass_count += 1

            if priority in {"High", "Critical"}:
                high_priority_count += 1

            rows.append({
                "ticket_id": ticket.get("id"),
                "item_code": ticket.get("itemCode"),
                "item_name": ticket.get("itemName"),
                "wo": ticket.get("wo"),
                "branch": ticket.get("branch"),
                "qty_order": ticket.get("qtyOrder"),
                "qty_quarantine": ticket.get("qtyQuarantine"),
                "defect_owner": ticket.get("defectOwner"),
                "where_detected": ticket.get("whereDetected"),
                "defect_code": ticket.get("defectCode"),
                "defect_description": ticket.get("defectDescription"),
                "corrective_action": ticket.get("correctiveAction"),
                "preventive_action": ticket.get("preventiveAction"),
                "priority": priority,
                "deadline": ticket.get("deadline"),
                "from_dept": ticket.get("fromDept"),
                "crew": ticket.get("crew"),
                "branch_route": ticket.get("branchRoute"),
                "supplement_needed": bool((ticket.get("supplement") or {}).get("needed")),
                "inspector": source.get("inspector"),
                "inspector_name": source.get("inspectorName"),
                "inspector_dept": source.get("inspectorDept"),
                "inspection_stage": source.get("inspectionStage"),
                "inspected_date": source.get("inspectedDate"),
                "pass_qty": pass_qty,
                "fail_qty": fail_qty,
                "defect_category": source.get("defectCategory"),
                "remark": source.get("remark"),
                "check_pass_date": source.get("checkPassDate"),
                "check_pass_minutes": source.get("checkPassMinutes"),
                "check_pass_hours": source.get("checkPassHours"),
                "status": "Fail" if fail_qty > 0 else "Pass"
            })

        return {
            "summary": {
                "total_reports": len(rows),
                "pass_reports": pass_count,
                "fail_reports": fail_count,
                "high_priority_reports": high_priority_count,
            },
            "reports": rows,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/qc-manager/dashboard")
def get_dashboard(
    request: Request,
    current_user: dict = Depends(get_current_user),
):
    """Get QC Manager dashboard data with quality metrics from the shared inspection store."""
    if current_user['role'] != 'QC Manager':
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        params = dict(request.query_params) if request else {}
        include_unknown = str(params.get('include_unknown', 'false')).lower() in {'1', 'true', 'yes'}
        stage_filter = params.get('stage') or params.get('inspection_stage') or None
        date_from = params.get('date_from') or None
        date_to = params.get('date_to') or None
        defect_code = params.get('defect_code') or None

        records = _load_store()
        if stage_filter:
            records = [record for record in records if _match_query_value(record.get('inspection_stage'), stage_filter)]
        if defect_code:
            records = [record for record in records if _match_query_value(record.get('defect_code'), defect_code)]
        if date_from:
            from_date = _parse_datetime(date_from)
            if from_date:
                records = [record for record in records if (_parse_datetime(record.get('inspected_date')) or _parse_datetime(record.get('inspected_time')) or datetime.min) >= from_date]
        if date_to:
            to_date = _parse_datetime(date_to)
            if to_date:
                records = [record for record in records if (_parse_datetime(record.get('inspected_date')) or _parse_datetime(record.get('inspected_time')) or datetime.min) <= to_date]

        return _compute_dashboard_payload(records, include_unknown=include_unknown)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/qc-manager/cost-management")
def get_cost_management(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get cost management data from supplement request approvals (QC Manager only)"""
    if current_user['role'] != 'QC Manager':
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        repairs = db.query(RepairCompletionDB).all()
        total_cost = 0
        cost_details = []
        
        for repair in repairs:
            if repair.actual_cost is not None:
                total_cost += repair.actual_cost
                cost_details.append({
                    "repair_id": repair.repair_completion_id,
                    "cost": repair.actual_cost,
                    "description": repair.root_cause,
                    "date": repair.completion_date.isoformat() if repair.completion_date else None
                })
        
        return {
            "total_cost": total_cost,
            "cost_details": cost_details,
            "total_repairs": len(cost_details)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/qc-manager/inspection-control")
def get_inspection_control(
    request: Request,
    current_user: dict = Depends(get_current_user), 
    limit: int = 50,
    offset: int = 0,
):
    """Get inspection control data with advanced filters from the shared file store."""
    if current_user['role'] != 'QC Manager':
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        params = dict(request.query_params) if request else {}
        excluded = {'limit', 'offset'}
        records = _apply_column_filters(_load_store(), params, excluded)
        total = len(records)
        page = records[offset:offset + limit]
        return {
            "inspections": [_to_inspection_control_row(record) for record in page],
            "total": total,
            "limit": limit,
            "offset": offset,
            "has_more": offset + limit < total,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/qc-manager/defect-frequency")
def get_defect_frequency(request: Request, current_user: dict = Depends(get_current_user)):
    """Get defect frequency for Quality Control dashboard from the shared file store."""
    if current_user['role'] != 'QC Manager':
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        params = dict(request.query_params) if request else {}
        stage_filter = params.get('stage') or params.get('inspection_stage') or None
        defect_code = params.get('defect_code') or None
        date_from = params.get('date_from') or None
        date_to = params.get('date_to') or None

        records = _load_store()
        if stage_filter:
            records = [record for record in records if _match_query_value(record.get('inspection_stage'), stage_filter)]
        if defect_code:
            records = [record for record in records if _match_query_value(record.get('defect_code'), defect_code)]
        if date_from:
            from_date = _parse_datetime(date_from)
            if from_date:
                records = [record for record in records if (_parse_datetime(record.get('inspected_date')) or _parse_datetime(record.get('inspected_time')) or datetime.min) >= from_date]
        if date_to:
            to_date = _parse_datetime(date_to)
            if to_date:
                records = [record for record in records if (_parse_datetime(record.get('inspected_date')) or _parse_datetime(record.get('inspected_time')) or datetime.min) <= to_date]

        frequency = {}
        
        for record in records:
            if not (_coerce_bool(record.get('quarantine')) or _coerce_int(record.get('fail_qty'), 0) > 0):
                continue
            code = record.get("defect_code") or "Unknown"
            if code not in frequency:
                frequency[code] = {"count": 0, "type": record.get("defect_category") or "Unknown"}
            frequency[code]["count"] += 1
        
        # Sort by frequency
        sorted_freq = sorted(frequency.items(), key=lambda x: x[1]["count"], reverse=True)
        
        return {
            "defect_frequency": [
                {"code": code, "type": data["type"], "count": data["count"]} 
                for code, data in sorted_freq
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/qc-manager/cost-breakdown")
def get_cost_breakdown(
    current_user: dict = Depends(get_current_user), 
    db: Session = Depends(get_db),
    group_by: Optional[str] = None
):
    """Get detailed cost breakdown grouped by stage/department/defect code"""
    if current_user['role'] != 'QC Manager':
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        repairs = db.query(RepairCompletionDB).all()
        costs = {}
        total_labor_cost = 0
        total_spare_cost = 0
        
        for repair in repairs:
            plan = db.query(RepairPlanDB).filter(RepairPlanDB.plan_id == repair.repair_plan_id).first()
            assignment = None
            inspection = None
            if plan:
                assignment = db.query(OwnerAssignmentDB).filter(OwnerAssignmentDB.assignment_id == plan.assignment_id).first()
                if assignment:
                    quarantine = db.query(QuarantineReportDB).filter(QuarantineReportDB.quarantine_id == assignment.quarantine_id).first()
                    if quarantine:
                        inspection = db.query(InspectionDB).filter(InspectionDB.inspection_id == quarantine.inspection_id).first()

            labor_cost = repair.labor_cost or 0
            spare_cost = repair.spare_cost or 0
            copq = labor_cost + spare_cost
            
            total_labor_cost += labor_cost
            total_spare_cost += spare_cost
            
            # Use group_by parameter to organize costs
            key = "Total"  # Default key
            if group_by == "stage" and inspection:
                key = inspection.inspection_stage or 'Unknown'
            elif group_by == "department" and assignment:
                key = str(assignment.repair_department_id or 'Unknown')
            
            if key not in costs:
                costs[key] = {"labor_cost": 0, "spare_cost": 0, "copq": 0, "count": 0}
            
            costs[key]["labor_cost"] += labor_cost
            costs[key]["spare_cost"] += spare_cost
            costs[key]["copq"] += copq
            costs[key]["count"] += 1
        
        return {
            "total_labor_cost": total_labor_cost,
            "total_spare_cost": total_spare_cost,
            "total_copq": total_labor_cost + total_spare_cost,
            "breakdown": costs
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============ PRODUCTION MANAGER ENDPOINTS ============

@app.get("/api/production-manager/notifications")
def get_notifications(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get notifications for Production Manager"""
    if current_user['role'] != 'Production Manager':
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        assignments = db.query(OwnerAssignmentDB).all()
        result = []
        
        for oa in assignments:
            quarantine = db.query(QuarantineReportDB).filter(QuarantineReportDB.quarantine_id == oa.quarantine_id).first()
            inspection = db.query(InspectionDB).filter(InspectionDB.inspection_id == quarantine.inspection_id).first() if quarantine else None
            
            result.append({
                "assignment_id": oa.assignment_id,
                "priority": oa.priority,
                "product_id": inspection.product_id if inspection else "N/A",
                "batch_number": inspection.batch_number if inspection else "N/A",
                "assigned_date": oa.assigned_date.isoformat() if oa.assigned_date else None
            })
        
        return {"notifications": result}
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/production-manager/repair-plan")
def create_repair_plan(plan: RepairPlan, current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Create repair plan (Production Manager)"""
    if current_user['role'] != 'Production Manager':
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        new_plan = RepairPlanDB(
            assignment_id=plan.assignment_id,
            plan_description=plan.plan_description,
            estimated_days=plan.estimated_days,
            assigned_worker_id=plan.assigned_worker_id
        )
        db.add(new_plan)
        db.commit()
        db.refresh(new_plan)
        
        return {"plan_id": new_plan.plan_id, "status": "created"}
    
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# ============ PRODUCTION WORKER ENDPOINTS ============

@app.get("/api/production-worker/repair-plans")
def get_repair_plans(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get assigned repair plans (Production Worker)"""
    if current_user['role'] not in ('Production Worker', 'Production Manager', 'QC Manager'):
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        plans = db.query(RepairPlanDB).filter(RepairPlanDB.assigned_worker_id == current_user['user_id']).all()
        
        result = []
        for rp in plans:
            barcode = "1234567890123"
            item_name = "Motor Housing"
            part_no = "ABC-001"
            model = "MH-100"
            
            oa = db.query(OwnerAssignmentDB).filter(OwnerAssignmentDB.assignment_id == rp.assignment_id).first()
            if oa:
                quarantine = db.query(QuarantineReportDB).filter(QuarantineReportDB.quarantine_id == oa.quarantine_id).first()
                if quarantine:
                    inspection = db.query(InspectionDB).filter(InspectionDB.inspection_id == quarantine.inspection_id).first()
                    if inspection:
                        barcode = inspection.product_id or "1234567890123"
                        item_name = inspection.component_code or "Wood Component"
                        part_no = inspection.carcass_code or "ABC-001"
                        model = "MH-100"
            
            result.append({
                "plan_id": rp.plan_id,
                "description": rp.plan_description,
                "estimated_days": rp.estimated_days,
                "status": rp.status,
                "created_date": rp.created_date.isoformat() if rp.created_date else None,
                "barcode": barcode,
                "item_name": item_name,
                "part_no": part_no,
                "model": model
            })
        
        return {"plans": result}
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/production-worker/repair-plans/{plan_id}/start")
def start_repair_plan(plan_id: int, current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Start an assigned repair plan and mark it in progress."""
    if current_user['role'] not in ('Production Worker', 'Production Manager', 'QC Manager'):
        raise HTTPException(status_code=403, detail="Unauthorized")

    try:
        plan = db.query(RepairPlanDB).filter(RepairPlanDB.plan_id == plan_id).first()
        if not plan:
            raise HTTPException(status_code=404, detail="Repair plan not found")

        if current_user['role'] == 'Production Worker' and plan.assigned_worker_id != current_user['user_id']:
            raise HTTPException(status_code=403, detail="Repair plan is assigned to another worker")

        if plan.status == "completed":
            raise HTTPException(status_code=400, detail="Completed repair plans cannot be restarted")

        plan.status = "in_progress"
        db.commit()
        db.refresh(plan)

        return {"plan_id": plan.plan_id, "status": plan.status}

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/production-worker/repair-completion")
def complete_repair(completion: RepairCompletion, current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Complete repair and record details (Production Worker)"""
    if current_user['role'] not in ('Production Worker', 'Production Manager', 'QC Manager'):
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        new_completion = RepairCompletionDB(
            repair_plan_id=completion.repair_plan_id,
            actual_cost=completion.actual_cost,
            completion_date=datetime.utcnow(),
            root_cause=completion.root_cause,
            preventive_action=completion.preventive_action
        )
        db.add(new_completion)
        db.commit()
        
        # Update repair plan status
        plan = db.query(RepairPlanDB).filter(RepairPlanDB.plan_id == completion.repair_plan_id).first()
        if plan:
            plan.status = "completed"
            db.commit()
        
        db.refresh(new_completion)
        
        return {"completion_id": new_completion.completion_id, "status": "completed"}
    
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# ============ QC MANAGER DEFECT CONTROL ENDPOINTS ============

@app.get("/api/qc-manager/defect-control/notifications")
def qcm_get_defect_notifications(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get notifications for QC Manager Defect Control (mirrors Production Manager notifications)"""
    if current_user['role'] not in ('QC Manager', 'Production Manager'):
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        assignments = db.query(OwnerAssignmentDB).all()
        result = []
        
        for oa in assignments:
            quarantine = db.query(QuarantineReportDB).filter(QuarantineReportDB.quarantine_id == oa.quarantine_id).first()
            inspection = db.query(InspectionDB).filter(InspectionDB.inspection_id == quarantine.inspection_id).first() if quarantine else None
            
            result.append({
                "assignment_id": oa.assignment_id,
                "priority": oa.priority,
                "product_id": inspection.product_id if inspection else "N/A",
                "batch_number": inspection.batch_number if inspection else "N/A",
                "assigned_date": oa.assigned_date.isoformat() if oa.assigned_date else None
            })
        
        return {"notifications": result}
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/qc-manager/defect-control/repair-plan")
def qcm_create_repair_plan(plan: RepairPlan, current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Create repair plan (QC Manager Defect Control — mirrors Production Manager repair-plan)"""
    if current_user['role'] not in ('QC Manager', 'Production Manager'):
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        new_plan = RepairPlanDB(
            assignment_id=plan.assignment_id,
            plan_description=plan.plan_description,
            estimated_days=plan.estimated_days,
            assigned_worker_id=plan.assigned_worker_id
        )
        db.add(new_plan)
        db.commit()
        db.refresh(new_plan)
        
        return {"plan_id": new_plan.plan_id, "status": "created"}
    
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# ============ WAREHOUSE ENDPOINTS ============

@app.get("/api/warehouse/repairs-for-approval")
def get_repairs_for_approval(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get repairs pending approval (Warehouse)"""
    if current_user['role'] != 'Warehouse':
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        completions = db.query(RepairCompletionDB).all()
        result = []
        
        for rc in completions:
            plan = db.query(RepairPlanDB).filter(RepairPlanDB.plan_id == rc.repair_plan_id).first()
            approval = db.query(ApprovalDB).filter(ApprovalDB.repair_plan_id == rc.repair_plan_id).first()
            
            if not approval or approval.status == "pending":
                result.append({
                    "completion_id": rc.completion_id,
                    "plan_id": rc.repair_plan_id,
                    "description": plan.plan_description if plan else "N/A",
                    "actual_cost": rc.actual_cost,
                    "root_cause": rc.root_cause,
                    "preventive_action": rc.preventive_action,
                    "completion_date": rc.completion_date.isoformat() if rc.completion_date else None
                })
        
        return {"repairs": result}
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/warehouse/approve-repair")
def approve_repair(approval: Approval, current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Approve or reject repair (Warehouse)"""
    if current_user['role'] != 'Warehouse':
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    try:
        new_approval = ApprovalDB(
            repair_plan_id=approval.repair_plan_id,
            status=approval.status,
            approver_id=current_user['user_id'],
            comments=approval.comments
        )
        db.add(new_approval)
        db.commit()
        db.refresh(new_approval)
        
        return {"approval_id": new_approval.approval_id, "status": approval.status}
    
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# ============ MOUNT STATIC FILES ============

def parse_excel_data(xlsx_path: Path) -> list:
    import openpyxl
    
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    
    headers = [cell.value for cell in worksheet[1]]
    tickets = []
    
    def _as_text(value) -> str:
        if value is None:
            return ""
        return str(value).strip()
        
    def _format_deadline(inspection_date, priority: str) -> str:
        if inspection_date is None:
            return "Pending"
        if priority == "High":
            deadline = inspection_date.replace(hour=17, minute=30, second=0, microsecond=0)
        elif priority == "Medium":
            deadline = inspection_date.replace(hour=19, minute=0, second=0, microsecond=0)
        else:
            deadline = (inspection_date + timedelta(days=1)).replace(hour=10, minute=0, second=0, microsecond=0)
        return deadline.strftime("%Y-%m-%d %H:%M")

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        record = dict(zip(headers, row))
        wo_number = record.get("WO No.")
        item_code = _as_text(record.get("Item Code"))
        part_code = _as_text(record.get("Partcode"))
        
        if not wo_number and not item_code and not part_code:
            continue
            
        inspection_date = record.get("Inspected Date")
        if isinstance(inspection_date, str):
            try:
                inspection_date = datetime.fromisoformat(inspection_date)
            except ValueError:
                inspection_date = None
                
        fail_qty = int(record.get("Fail Qty") or 0)
        if fail_qty <= 0:
            continue
        defect_category = _as_text(record.get("Defect Category"))
        priority = "High" if defect_category == "Major" or fail_qty >= 1 else "Medium"
        defect_code = _as_text(record.get("Defect Code"))
        defect_description = defect_code or "Unknown defect"
        corrective_action = _as_text(record.get("Inspector Recommend")) or "Review and rework"
        repair_department = _as_text(record.get("Repair Department")) or "Unknown"
        inspector_name = _as_text(record.get("Inspector Name"))
        inspector_dept = _as_text(record.get("Inspector Dept."))
        stage = _as_text(record.get("Inspection Stage"))
        
        tickets.append({
            "id": f"INH-{row_index}-{wo_number or item_code or part_code}",
            "itemCode": item_code or part_code or f"WO-{wo_number}",
            "itemName": part_code or item_code or f"WO {wo_number}",
            "wo": _as_text(wo_number),
            "customer": "",
            "branch": _as_text(record.get("Branch ITW")) or _as_text(record.get("Branch FGW")),
            "qtyOrder": int(record.get("WO Quantity") or 0),
            "qtyQuarantine": int(record.get("Quarantine") or fail_qty or 0),
            "defectOwner": _as_text(record.get("Defect Owner")),
            "whereDetected": f"{stage} — {inspector_dept}" if stage or inspector_dept else stage or inspector_dept,
            "defectCode": defect_code,
            "defectDescription": defect_description,
            "correctiveAction": corrective_action,
            "preventiveAction": "",
            "priority": priority,
            "deadline": _format_deadline(inspection_date, priority),
            "fromDept": repair_department,
            "crew": "1 worker" if fail_qty <= 1 else f"{fail_qty} workers",
            "branchRoute": "yes" if repair_department else "no",
            "supplement": {
                "needed": False
            },
            "source": {
                "inspector": _as_text(record.get("Inspector")),
                "inspectorName": inspector_name,
                "inspectorDept": inspector_dept,
                "inspectionStage": stage,
                "inspectedDate": inspection_date.isoformat() if isinstance(inspection_date, datetime) else _as_text(record.get("Inspected Date")),
                "passQty": int(record.get("Pass Qty") or 0),
                "failQty": fail_qty,
                "defectCategory": defect_category,
                "remark": _as_text(record.get("Remark")),
                "checkPassDate": _as_text(record.get("Check Pass Date")),
                "checkPassMinutes": _as_text(record.get("Check Pass Minutes")),
                "checkPassHours": _as_text(record.get("Check Pass Hours")),
            }
        })
    return tickets

def ensure_inspection_data():
    json_path = Path(__file__).resolve().parent / "public" / "inspection-data.json"
    xlsx_path = Path(__file__).resolve().parent / "data" / "input data.xlsx"
    
    if xlsx_path.exists():
        should_generate = False
        if not json_path.exists():
            should_generate = True
        else:
            if xlsx_path.stat().st_mtime > json_path.stat().st_mtime:
                should_generate = True
                
        if should_generate:
            print("Auto-generating public/inspection-data.json from data/input data.xlsx...")
            try:
                tickets = parse_excel_data(xlsx_path)
                json_path.parent.mkdir(parents=True, exist_ok=True)
                json_path.write_text(json.dumps(tickets, ensure_ascii=False, indent=2), encoding="utf-8")
                print(f"Successfully generated: {len(tickets)} records.")
            except Exception as e:
                print(f"Failed to auto-generate JSON from Excel: {e}")

@app.get("/data/input data")
def get_input_data_custom():
    """Return the imported fail-file data at custom path."""
    ensure_inspection_data()
    data_file = Path(__file__).resolve().parent / "public" / "inspection-data.json"
    if not data_file.exists():
        raise HTTPException(status_code=404, detail="inspection-data.json not found")
    try:
        return json.loads(data_file.read_text(encoding="utf-8"))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/inspection-data")
def get_inspection_data():
    """Return the imported fail-file data used by Production Worker."""
    ensure_inspection_data()
    data_file = Path(__file__).resolve().parent / "public" / "inspection-data.json"
    if not data_file.exists():
        raise HTTPException(status_code=404, detail="inspection-data.json not found")

    try:
        return json.loads(data_file.read_text(encoding="utf-8"))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

_last_excel_sync_time = 0.0

def sync_to_db(db: Session, card: dict, request_code: str, status: str):
    card_id = str(card.get("id"))
    wo_no = str(card.get("wo_no") or card.get("wo") or "").strip()
    item_code = str(card.get("item_code") or card.get("itemCode") or "").strip()
    carcass_code = str(card.get("carcass_code") or card.get("partcode") or "").strip()
    
    # 1. Update/Insert in repair_card_supplement_link
    link = db.query(RepairCardSupplementLinkDB).filter(
        RepairCardSupplementLinkDB.repair_card_id == card_id,
        RepairCardSupplementLinkDB.supplement_request_code == request_code
    ).first()
    
    if not link:
        link = RepairCardSupplementLinkDB(
            repair_card_id=card_id,
            supplement_request_code=request_code,
            wo_no=wo_no,
            item_code=item_code,
            carcass_code=carcass_code
        )
        db.add(link)
    
    # 2. Update InspectionDB
    inspection = db.query(InspectionDB).filter(
        InspectionDB.wo_number == wo_no,
        InspectionDB.batch_number == item_code
    ).first()
    if inspection:
        inspection.repair_status = status
        inspection.repair_status_updated_at = datetime.utcnow()

def sync_repair_cards_and_supplement_requests(db: Optional[Session] = None, force: bool = False):
    global _last_excel_sync_time
    
    supp_path = Path("data/Supplement request.xlsx")
    link_path = Path("data/Sample_Repair_Supplement_Link.xlsx")
    
    mtime1 = supp_path.stat().st_mtime if supp_path.exists() else 0.0
    mtime2 = link_path.stat().st_mtime if link_path.exists() else 0.0
    current_max_mtime = max(mtime1, mtime2)
    
    if not force and current_max_mtime <= _last_excel_sync_time:
        return
        
    supplement_items = []
    
    if supp_path.exists():
        try:
            workbook = load_workbook(supp_path, data_only=True)
            for sheet_name in workbook.sheetnames:
                if sheet_name.startswith("SRC-"):
                    sheet = workbook[sheet_name]
                    rows = list(sheet.iter_rows(values_only=True))
                    if len(rows) > 1:
                        headers = [str(h).strip() if h is not None else f"Col{idx}" for idx, h in enumerate(rows[0])]
                        for row in rows[1:]:
                            if all(cell is None for cell in row):
                                continue
                            item = dict(zip(headers, row))
                            wo = item.get("WO") or item.get("WO No.")
                            item_code = item.get("Item Code")
                            if wo and item_code:
                                supplement_items.append({
                                    "request_code": sheet_name,
                                    "wo": str(wo).strip(),
                                    "item_code": str(item_code).strip(),
                                    "carcass_code": str(item.get("Carcass Code") or "").strip(),
                                    "material_code": str(item.get("Material Code") or "").strip(),
                                    "material_name": str(item.get("Material Name") or "").strip(),
                                    "qty": item.get("Supplement Qty") or 0,
                                    "hod": str(item.get("HOD") or "").strip(),
                                    "related_hod": str(item.get("Related HOD") or "").strip(),
                                    "bod": str(item.get("BOD") or "").strip()
                                })
        except Exception as e:
            print(f"Error loading Supplement request.xlsx: {e}")

    if link_path.exists():
        try:
            workbook = load_workbook(link_path, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) > 1:
                headers = [str(h).strip() if h is not None else f"Col{idx}" for idx, h in enumerate(rows[0])]
                for row in rows[1:]:
                    if all(cell is None for cell in row):
                        continue
                    item = dict(zip(headers, row))
                    wo = item.get("WO No.")
                    item_code = item.get("Item Code")
                    req_code = item.get("Supplement Request Code")
                    if wo and item_code and req_code:
                        supplement_items.append({
                            "request_code": str(req_code).strip(),
                            "wo": str(wo).strip(),
                            "item_code": str(item_code).strip(),
                            "carcass_code": str(item.get("Carcass Code") or "").strip(),
                            "material_code": str(item.get("Material Code") or "").strip(),
                            "material_name": str(item.get("Material Name") or "").strip(),
                            "qty": item.get("Supplement Qty") or 0,
                            "hod": str(item.get("HOD Approval") or "").strip(),
                            "related_hod": str(item.get("Related HOD Approval") or "").strip(),
                            "bod": str(item.get("BOD Approval") or "").strip()
                        })
        except Exception as e:
            print(f"Error loading Sample_Repair_Supplement_Link.xlsx: {e}")

    if not supplement_items:
        return

    # Group supplement items by (wo, item_code)
    supplement_map = {}
    for item in supplement_items:
        wo = item["wo"]
        item_code = item["item_code"]
        key = (wo, item_code)
        if key not in supplement_map:
            supplement_map[key] = []
        supplement_map[key].append(item)

    json_paths = [
        Path("data/inspection-control.json"),
        Path("public/inspection-data.json")
    ]
    
    local_db = db or SessionLocal()
    try:
        for json_path in json_paths:
            if not json_path.exists():
                continue
            try:
                cards = json.loads(json_path.read_text(encoding="utf-8"))
                changed = False
                for card in cards:
                    wo = str(card.get("wo_no") or card.get("wo") or "").strip()
                    item_code = str(card.get("item_code") or card.get("itemCode") or "").strip()
                    carcass_code = str(card.get("carcass_code") or "").strip()
                    
                    if not wo or not item_code:
                        continue
                    
                    key = (wo, item_code)
                    matches = supplement_map.get(key) or []
                    
                    # Match on carcass_code if present on both sides
                    filtered_matches = []
                    for m in matches:
                        req_carcass = m["carcass_code"]
                        if req_carcass and carcass_code:
                            if req_carcass.lower() in carcass_code.lower() or carcass_code.lower() in req_carcass.lower():
                                filtered_matches.append(m)
                        else:
                            filtered_matches.append(m)
                    
                    if filtered_matches:
                        req_code = filtered_matches[0]["request_code"]
                        
                        all_approved = True
                        for m in filtered_matches:
                            hod_ok = m["hod"].lower() == "approved"
                            rel_ok = m["related_hod"].lower() == "approved"
                            bod_ok = m["bod"].lower() == "approved"
                            if not (hod_ok and rel_ok and bod_ok):
                                all_approved = False
                                break
                        
                        target_status = "IN_REPAIR" if all_approved else "WAITING_MATERIAL"
                        current_status = card.get("repair_status")
                        
                        if current_status != target_status or card.get("supplement_request_code") != req_code:
                            card["repair_status"] = target_status
                            card["repair_status_updated_at"] = datetime.utcnow().isoformat()
                            card["supplement_request_code"] = req_code
                            
                            # Update supplement dictionary
                            card["supplement"] = {
                                "needed": True,
                                "status": "waiting" if target_status == "WAITING_MATERIAL" else "arrived",
                                "partName": filtered_matches[0]["material_name"] or "Material Supplement",
                                "qty": filtered_matches[0]["qty"] or 1,
                                "eta": "Pending" if target_status == "WAITING_MATERIAL" else "Storage Bin",
                                "requestCode": req_code
                            }
                            changed = True
                            
                            try:
                                sync_to_db(local_db, card, req_code, target_status)
                            except Exception as dberr:
                                print(f"Error syncing to DB: {dberr}")
                                
                if changed:
                    json_path.write_text(json.dumps(cards, ensure_ascii=False, indent=2), encoding="utf-8")
                    print(f"Sync complete: {json_path.name} updated.")
            except Exception as e:
                print(f"Error syncing {json_path.name}: {e}")
        
        local_db.commit()
    finally:
        if db is None:
            local_db.close()
            
    _last_excel_sync_time = current_max_mtime

@app.post("/api/qc-manager/sync-repair-supplement")
def trigger_manual_sync(db: Session = Depends(get_db)):
    """Manually trigger supplement-request and repair-card sync."""
    try:
        sync_repair_cards_and_supplement_requests(db, force=True)
        return {"status": "success", "message": "Synchronization complete."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/qc-manager/seed-test-data")
def seed_test_data(db: Session = Depends(get_db)):
    """Seed test data from Sample_Repair_Supplement_Link.xlsx into the datastore."""
    try:
        link_path = Path("data/Sample_Repair_Supplement_Link.xlsx")
        if not link_path.exists():
            raise HTTPException(status_code=404, detail="Sample_Repair_Supplement_Link.xlsx not found")
            
        workbook = load_workbook(link_path, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) <= 1:
            raise HTTPException(status_code=400, detail="Sheet has no data rows")
            
        headers = [str(h).strip() if h is not None else f"Col{idx}" for idx, h in enumerate(rows[0])]
        
        json_paths = [
            Path("data/inspection-control.json"),
            Path("public/inspection-data.json")
        ]
        
        seeded_count = 0
        for r_idx, row in enumerate(rows[1:], start=2):
            if all(cell is None for cell in row):
                continue
            item = dict(zip(headers, row))
            wo = str(item.get("WO No.")).strip()
            item_code = str(item.get("Item Code")).strip()
            if not wo or not item_code:
                continue
                
            inspected_date = item.get("Inspected Date")
            if isinstance(inspected_date, datetime):
                inspected_date_str = inspected_date.isoformat()
            else:
                inspected_date_str = str(inspected_date) if inspected_date else datetime.utcnow().isoformat()
                
            card_id = f"INH-{r_idx}-{wo}"
            new_card = {
                "id": card_id,
                "branch_itw": str(item.get("Branch") or "VFR2.1").strip(),
                "branch_fgw": str(item.get("Branch") or "VFR2.1").strip(),
                "wo_no": wo,
                "item_code": item_code,
                "partcode": str(item.get("Partcode") or "").strip(),
                "carcass_code": str(item.get("Carcass Code") or "").strip(),
                "wo_quantity": int(item.get("WO Quantity") or 0),
                "inspector": "17372",
                "inspector_name": str(item.get("Inspector Name") or "Lê Thị Xuân").strip(),
                "inspector_dept": str(item.get("Inspector Dept.") or "2QCD").strip(),
                "inspection_stage": str(item.get("Inspection Stage") or "Finishing").strip(),
                "inspected_date": inspected_date_str,
                "inspected_time": inspected_date_str,
                "inspected_qty": 1,
                "pass_qty": 0,
                "fail_qty": int(item.get("Fail Qty") or 1),
                "quarantine": 1,
                "defect_code": str(item.get("Defect Code") or "").strip(),
                "defect_category": str(item.get("Defect Category") or "Major").strip(),
                "defect_owner": str(item.get("Defect Owner") or "").strip(),
                "inspector_recommend": str(item.get("Inspector Recommend") or "Sửa hàng - kiểm lại").strip(),
                "repair_department": str(item.get("Repair Department") or "").strip(),
                "remark": str(item.get("Trigger Note") or "").strip(),
                "check_pass_date": None,
                "check_pass_minutes": "",
                "check_pass_hours": "",
                "status": "Fail",
                "repair_status": "NEW",
                "repair_status_updated_at": None,
                "supplement_request_code": ""
            }
            
            for json_path in json_paths:
                cards = []
                if json_path.exists():
                    try:
                        cards = json.loads(json_path.read_text(encoding="utf-8"))
                    except Exception:
                        cards = []
                cards = [c for c in cards if c.get("id") != card_id]
                cards.append(new_card)
                json_path.write_text(json.dumps(cards, ensure_ascii=False, indent=2), encoding="utf-8")
                
            existing_inspection = db.query(InspectionDB).filter(
                InspectionDB.wo_number == wo,
                InspectionDB.batch_number == item_code
            ).first()
            
            if not existing_inspection:
                new_db_inspection = InspectionDB(
                    product_id=wo,
                    batch_number=item_code,
                    wo_number=wo,
                    component_code=item_code,
                    carcass_code=new_card["carcass_code"],
                    inspection_stage=new_card["inspection_stage"],
                    inspector_id=1,
                    qty_inspected=1,
                    qty_passed=0,
                    qty_failed=new_card["fail_qty"],
                    status="Fail",
                    repair_status="NEW"
                )
                db.add(new_db_inspection)
            
            seeded_count += 1
            
        db.commit()
        sync_repair_cards_and_supplement_requests(db, force=True)
        return {"status": "success", "message": f"Seeded {seeded_count} test records and synced status."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/supplement-requests")
def get_supplement_requests():
    """Return list of supplement requests parsed from Excel."""
    try:
        xlsx_path = Path("data/Supplement request.xlsx")
        if not xlsx_path.exists():
            raise HTTPException(status_code=404, detail="Supplement request.xlsx not found")
        
        workbook = load_workbook(xlsx_path, data_only=True)
        sheet = workbook["total"] if "total" in workbook.sheetnames else workbook.worksheets[0]
        
        requests = []
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if r_idx == 0:
                continue
            code = row[1]
            if code and str(code).startswith("SRC-"):
                create_date = row[6]
                if isinstance(create_date, datetime):
                    create_date = create_date.strftime("%d/%m/%Y")
                elif create_date:
                    try:
                        dt = datetime.fromisoformat(str(create_date))
                        create_date = dt.strftime("%d/%m/%Y")
                    except Exception:
                        pass
                
                expired_date = row[7]
                if isinstance(expired_date, datetime):
                    expired_date = expired_date.strftime("%d/%m/%Y")
                elif expired_date:
                    try:
                        dt = datetime.fromisoformat(str(expired_date))
                        expired_date = dt.strftime("%d/%m/%Y")
                    except Exception:
                        pass
                else:
                    expired_date = ""

                requests.append({
                    "requestCode": str(code).strip(),
                    "requestDept": str(row[2]).strip() if row[2] else "",
                    "relatedDept": str(row[3]).strip() if row[3] else "",
                    "status": str(row[4]).strip() if row[4] else "",
                    "createBy": str(row[5]).strip() if row[5] else "",
                    "createDate": create_date if create_date else "",
                    "expiredDate": expired_date
                })
                if len(requests) >= 3:
                    break
        return requests
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/supplement-requests/{request_code}")
def get_supplement_request_detail(request_code: str):
    """Return detailed items and master info for a specific supplement request."""
    try:
        xlsx_path = Path("data/Supplement request.xlsx")
        if not xlsx_path.exists():
            raise HTTPException(status_code=404, detail="Supplement request.xlsx not found")
        
        workbook = load_workbook(xlsx_path, data_only=True)
        if request_code not in workbook.sheetnames:
            raise HTTPException(status_code=404, detail=f"Request details for {request_code} not found")
            
        sheet = workbook[request_code]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return {"headers": [], "items": [], "master": {}}
            
        headers = [str(h).strip() if h is not None else f"Col{idx}" for idx, h in enumerate(rows[0])]
        
        items = []
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue
            if len(row) <= 2 or row[2] is None:
                continue
                
            item = {}
            for c_idx, h in enumerate(headers):
                if c_idx < len(row):
                    val = row[c_idx]
                    if isinstance(val, datetime):
                        val = val.strftime("%d/%m/%Y %H:%M:%S")
                    elif val is None:
                        val = ""
                    item[h] = val
                else:
                    item[h] = ""
            items.append(item)
            
        total_sheet = workbook["total"] if "total" in workbook.sheetnames else workbook.worksheets[0]
        master = {}
        for row in total_sheet.iter_rows(values_only=True):
            if row[1] and str(row[1]).strip() == request_code:
                create_date = row[6]
                if isinstance(create_date, datetime):
                    create_date = create_date.strftime("%d/%m/%Y")
                elif create_date:
                    try:
                        dt = datetime.fromisoformat(str(create_date))
                        create_date = dt.strftime("%d/%m/%Y")
                    except Exception:
                        pass
                
                master = {
                    "requestCode": request_code,
                    "requestDept": str(row[2]).strip() if row[2] else "",
                    "relatedDept": str(row[3]).strip() if row[3] else "",
                    "status": str(row[4]).strip() if row[4] else "",
                    "createBy": str(row[5]).strip() if row[5] else "",
                    "createDate": create_date if create_date else "",
                }
                break
                
        if request_code == "SRC-202606-0071":
            master.update({
                "reasonToRequest": "Construction change - Thay đổi thiết kế",
                "requester": "18679 - Phạm Thị Thu Uyên",
                "requesterRemark": "Tec updated drawing to improve the product's aesthetics",
                "hodApprovedBy": "17972 - Nguyễn Hoàng Tú",
                "hodApprovedDate": "29/06/2026",
                "hodRemark": "",
                "relatedHodApprovedBy": "16542 - Phạm Phú Quốc",
                "relatedHodApprovedDate": "29/06/2026",
                "relatedHodRemark": "Confirmed",
                "bodApprovedBy": "16848 - Jakob Svendsen",
                "bodApprovedDate": "30/06/2026",
                "bodRemark": "Confirmed"
            })
        elif request_code == "SRC-202607-0017":
            master.update({
                "reasonToRequest": "Design modification - Điều chỉnh kích thước",
                "requester": "18679 - Phạm Thị Thu Uyên",
                "requesterRemark": "Adjust dimension to fit layout design",
                "hodApprovedBy": "17972 - Nguyễn Hoàng Tú",
                "hodApprovedDate": "04/07/2026",
                "hodRemark": "",
                "relatedHodApprovedBy": "16542 - Phạm Phú Quốc",
                "relatedHodApprovedDate": "04/07/2026",
                "relatedHodRemark": "Confirmed",
                "bodApprovedBy": "16848 - Jakob Svendsen",
                "bodApprovedDate": "",
                "bodRemark": "Pending"
            })
        elif request_code == "SRC-202607-0010":
            master.update({
                "reasonToRequest": "Material shortage - Thiếu hụt nguyên vật liệu",
                "requester": "18679 - Phạm Thị Thu Uyên",
                "requesterRemark": "Supplement materials for urgent order",
                "hodApprovedBy": "17972 - Nguyễn Hoàng Tú",
                "hodApprovedDate": "03/07/2026",
                "hodRemark": "",
                "relatedHodApprovedBy": "16542 - Phạm Phú Quốc",
                "relatedHodApprovedDate": "03/07/2026",
                "relatedHodRemark": "Confirmed",
                "bodApprovedBy": "16848 - Jakob Svendsen",
                "bodApprovedDate": "",
                "bodRemark": "Pending"
            })
        else:
            master.update({
                "reasonToRequest": "General Supplement Request",
                "requester": "18679 - Phạm Thị Thu Uyên",
                "requesterRemark": "",
                "hodApprovedBy": "17972 - Nguyễn Hoàng Tú",
                "hodApprovedDate": "",
                "hodRemark": "",
                "relatedHodApprovedBy": "16542 - Phạm Phú Quốc",
                "relatedHodApprovedDate": "",
                "relatedHodRemark": "",
                "bodApprovedBy": "16848 - Jakob Svendsen",
                "bodApprovedDate": "",
                "bodRemark": ""
            })
            
        return {
            "headers": headers,
            "items": items,
            "master": master
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

app.mount("/", StaticFiles(directory="public", html=True), name="static")

# ============ RUN SERVER ============

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=False)



