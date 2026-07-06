#!/usr/bin/env python3
"""Convert the inhouse fail Excel report into the JSON format used by the UI."""

from __future__ import annotations

from datetime import datetime, timedelta
import json
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
SOURCE_FILE = ROOT / "data" / "input data.xlsx"
OUTPUT_FILE = ROOT / "public" / "inspection-data.json"


def _format_deadline(inspection_date: datetime | None, priority: str) -> str:
    if inspection_date is None:
        return "Pending"

    if priority == "High":
        deadline = inspection_date.replace(hour=17, minute=30, second=0, microsecond=0)
    elif priority == "Medium":
        deadline = inspection_date.replace(hour=19, minute=0, second=0, microsecond=0)
    else:
        deadline = (inspection_date + timedelta(days=1)).replace(hour=10, minute=0, second=0, microsecond=0)

    return deadline.strftime("%Y-%m-%d %H:%M")


def _as_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def convert_workbook_to_tickets() -> list[dict]:
    workbook = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    headers = [cell.value for cell in worksheet[1]]
    tickets: list[dict] = []

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        record = dict(zip(headers, row))
        wo_number = record.get("WO No.")
        item_code = _as_text(record.get("Item Code"))
        part_code = _as_text(record.get("Partcode"))

        if not wo_number and not item_code and not part_code:
            continue

        fail_qty = int(record.get("Fail Qty") or 0)
        if fail_qty <= 0:
            continue

        inspection_date = record.get("Inspected Date")
        if isinstance(inspection_date, str):
            try:
                inspection_date = datetime.fromisoformat(inspection_date)
            except ValueError:
                inspection_date = None

        fail_qty = int(record.get("Fail Qty") or 0)
        defect_category = _as_text(record.get("Defect Category"))
        priority = "High" if defect_category == "Major" or fail_qty >= 1 else "Medium"
        defect_code = _as_text(record.get("Defect Code"))
        defect_description = defect_code or "Unknown defect"
        corrective_action = _as_text(record.get("Inspector Recommend")) or "Review and rework"
        repair_department = _as_text(record.get("Repair Department")) or "Unknown"
        inspector_name = _as_text(record.get("Inspector Name"))
        inspector_dept = _as_text(record.get("Inspector Dept."))
        stage = _as_text(record.get("Inspection Stage"))

        tickets.append(
            {
                "id": f"INH-{row_index}-{wo_number or item_code or part_code}",
                "itemCode": item_code or part_code or f"WO-{wo_number}",
                "itemName": part_code or item_code or f"WO {wo_number}",
                "wo": _as_text(wo_number),
                "customer": "",
                "branch": _as_text(record.get("Branch ITW")) or _as_text(record.get("Branch FGW")),
                "qtyOrder": int(record.get("WO Quantity") or 0),
                "qtyQuarantine": int(record.get("Quarantine") or fail_qty or 0),
                "defectOwner": _as_text(record.get("Defect Owner")),
                "whereDetected": f"{stage} — {inspector_dept}" if stage or inspector_dept else stage or inspector_dept,
                "defectCode": defect_code,
                "defectDescription": defect_description,
                "correctiveAction": corrective_action,
                "preventiveAction": "",
                "priority": priority,
                "deadline": _format_deadline(inspection_date, priority),
                "fromDept": repair_department,
                "crew": "1 worker" if fail_qty <= 1 else f"{fail_qty} workers",
                "branchRoute": "yes" if repair_department else "no",
                "supplement": {
                    "needed": False
                },
                "source": {
                    "inspector": _as_text(record.get("Inspector")),
                    "inspectorName": inspector_name,
                    "inspectorDept": inspector_dept,
                    "inspectionStage": stage,
                    "inspectedDate": inspection_date.isoformat() if isinstance(inspection_date, datetime) else _as_text(record.get("Inspected Date")),
                    "passQty": int(record.get("Pass Qty") or 0),
                    "failQty": fail_qty,
                    "defectCategory": defect_category,
                    "remark": _as_text(record.get("Remark")),
                    "checkPassDate": _as_text(record.get("Check Pass Date")),
                    "checkPassMinutes": _as_text(record.get("Check Pass Minutes")),
                    "checkPassHours": _as_text(record.get("Check Pass Hours")),
                },
            }
        )

    return tickets


def main() -> None:
    if not SOURCE_FILE.exists():
        raise FileNotFoundError(f"Source workbook not found: {SOURCE_FILE}")

    tickets = convert_workbook_to_tickets()
    OUTPUT_FILE.write_text(json.dumps(tickets, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Imported {len(tickets)} records from {SOURCE_FILE.name} -> {OUTPUT_FILE.relative_to(ROOT)}")


if __name__ == "__main__":
    main()